import openpyxl

wb = openpyxl.load_workbook('Balance.xlsx')

print(wb.sheetnames)

ws = wb['Score']
print(ws)
ws1 = wb['Sheet1']
print(ws1)

wb.create_sheet('NewSheet')
wb.save('Balance.xlsx')